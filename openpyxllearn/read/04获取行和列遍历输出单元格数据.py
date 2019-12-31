# -*- coding:utf-8 -*-
import openpyxl


def main():
    """获取工作表"""
    # 读取一个xlsx格式文件（格式有严格要求）
    # 通过调用方法load_workbook(filename)进行文件读取，该方法中还有一个read_only参数用于设置文件打开方式，默认为可读可写，该方法最终将返回一个workbook的数据对象
    wb = openpyxl.load_workbook('工作簿.xlsx')
    # 获取所有工作表名（返回值形式为列表）
    sheets = wb.get_sheet_names()
    print(sheets)
    # 根据工作表名获得某一具体的工作表
    sheet_one = wb.get_sheet_by_name('Sheet1')
    print(sheet_one)
    # 获取工作表的表名,可以在任何时候通过Worksheet.title属性修改工作表名
    print(sheet_one.title)
    # 一般来说，表格大多数用到的是打开时显示的工作表，这时可以用active来获取当前工作表
    # 工作表在工作簿创建后,可以通过Workbook.active属性来定位到工作表
    sheet_one = wb.active
    # 对行进行遍历,输出A1,B1,C1
    """
       在处理Excel表格有时可能需要对表格进行遍历查找，openpyxl中便提供了一个行和列的生成器(sheet.rows和sheet.columns),
       这两个生成器里面是每一行（或列）的数据，每一行（或列）又由一个tuple包裹，借此可以很方便地完成对行和列的遍历
    """
    # sheet.rows(或sheet.columns)是生成器类型，是不能直接调用的，需将其转化为一个list类型，然后再通过索引遍历
    for row in list(sheet_one.rows)[0]:
        for cell in row:
            print(cell.value)

    # 对列进行遍历,输出A1,A2,A3
    for column in list(sheet_one.columns)[0]:
        for cell in column:
            print(cell.value)
    # 对单元格的范围进行遍历 通过使用sheet[行列值:行列值]来对给定单元格范围进行遍历
    for spaces in sheet_one['A1':'B2']:
        for cell in spaces:
            print(cell.value)
    # 获取最大列
    print(sheet_one.max_column)
    # 获取最大行
    print(sheet_one.max_row)


if __name__ == '__main__':
    main()