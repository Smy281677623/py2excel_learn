# -*- coding:utf-8 -*-
import openpyxl


def main():
    # 打开一个excel
    wb = openpyxl.load_workbook('工作簿.xlsx')
    # 获得一个具体的sheet
    sheet1 = wb.get_sheet_by_name('sheet1')
    # 通过sheet[列行名]获取 单元格对象
    a = sheet1['A2']
    print(a)
    # 通过sheet.cell(row,column)获取 单元格对象
    # 需要注意的是，sheet.cell(row,column)中参数分别是行和列，且必须为整数，如果列为英文字母，
    # 可以利用 openpyxl.utils 中的   column_index_from_string(char)进行字母数字的转化。
    # 顺便一说，同理也可以利用get_column_letter(number)进行数字字母间的转化
    b = sheet1.cell(1, 2) # B1
    # 获取单元格内容
    print(a.value)
    print(b.value)
    # 获取单元格所在列和行
    print('a is '+str((a.column, a.row)))




if __name__ == '__main__':
    main()