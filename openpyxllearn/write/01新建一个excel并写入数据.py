# -*- coding:utf-8 -*-
import openpyxl

def main():
    # 创建一个工作簿对象
    wb = openpyxl.Workbook()
    # 新建的工作簿会自动设置一个sheet表
    # 删除默认的sheet
    # wb.remove('sheet')
    # 创建我们自己的sheet create_sheet(工作表名, 索引名[从0开始])
    wb.create_sheet('1', 0)
    # get_sheet_by_name('工作表名称') 获得我们自己创建的sheet对象
    # sheet_one = wb.get_sheet_by_name('1')
    # 可以利用  工作簿对象['工作表名称'] 数组索引的方式获取工作表对象
    sheet_one = wb['1']
    # 创建的工作表的标签背景色默认是白色。可以通过在Worksheet.sheet_properties.tabColor对象中设
    sheet_one.sheet_properties.tabColor = "1072BA"
    # 写入单元格方式一 直接赋值
    sheet_one['A1'].value = 2
    # 写入单元格方式二 公式赋值
    sheet_one['A6'].value = '=SUM(A1:A5)'
    # 写入一行数据
    row = [1, 2, 3, 4, 5]
    sheet_one.append(row)
    # 写入多行数据
    # 写入多行
    rows = [
        ['ID'.encode('utf-8'), 'Name'.encode('utf-8'), 'Department'.encode('utf-8')],
        ['001'.encode('utf-8'), 'Lee'.encode('utf-8'), 'CS'.encode('utf-8')],
        ['002'.encode('utf-8'), 'John'.encode('utf-8'), 'MA'.encode('utf-8')],
        ['003'.encode('utf-8'), 'Amy'.encode('utf-8'), 'IS'.encode('utf-8')]
    ]
    for temp in rows:
        sheet_one.append(temp)
    # 最后不要忘记保存文件,文件格式必须是xlsx格式
    wb.save('openpyxlCreateOne.xlsx')


if __name__ == '__main__':
    main()