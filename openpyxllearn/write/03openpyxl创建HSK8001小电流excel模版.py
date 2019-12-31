# -*- coding:utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
import time
# 4个数据集合 传进封装类 实体

class Excel_Template(object):
    def __init__(self, board_voltage, mu_voltage, board_current, mu_current, board_num):
        self.board_voltage_ = board_voltage
        self.mu_voltage_ = mu_voltage
        self.board_current_ = board_current
        self.mu_current_ = mu_current
        self.board_num_ = board_num
        self.setup_excel()

    def setup_excel(self):
        """创建excel模版"""
        wb = openpyxl.Workbook()
        # 如果不给sheet名称则直接创建默认的sheet1名称
        if self.board_num_ is None:
            ws = wb.create_sheet('sheet1', 0)
        else:
            ws = wb.create_sheet(str(self.board_num_), 0)
        # 居中格式设置
        ali = Alignment(horizontal='center', vertical='center')
        # 边框设置
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        myborder = Border(left=left, right=right, top=top, bottom=bottom)

        # 5种颜色配置
        SpringGreen2 = PatternFill("solid", fgColor="6A5ACD")
        DeepSkyBlue = PatternFill("solid", fgColor="00BFFF")
        Turquoise = PatternFill("solid", fgColor="40E0D0")
        Yellow3 = PatternFill("solid", fgColor="CDCD00")
        PaleGoldenrod = PatternFill("solid", fgColor="EEE8AA")
        white = PatternFill("solid", fgColor="FFFFFF")

        # 设置 1 2 3 4 5 6 列的宽度
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20.0
        # 设置小电流模式
        for t in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']:
            ws[t].alignment = ali
            ws[t].border = myborder
            ws[t].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)

        ws['A1'].value = '小电流模式'
        ws['B1'].value = '机器编码'

        ws.merge_cells('B1:C1')
        ws['D1'].value = '板卡条码'
        ws.merge_cells('D1:F1')

        ws['A2'].value = '设置电压'
        ws['B2'].value = '小电流模式采样次数'
        ws['C2'].value = '8001电压读数（mV）'
        ws['D2'].value = '万用表电压读数（mV）'
        ws['E2'].value = '8001电流读数(mA)'
        ws['F2'].value = '万用表电流读数(mA)'

        for t in ['A2', 'B2', 'C2', 'D2', 'E2', 'F2']:
            ws[t].alignment = ali
            ws[t].border = myborder
            ws[t].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)

        # 第一轮老化测试
        ws['A3'].value = '第一轮老化测试'
        ws['A3'].alignment = ali
        ws['A3'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in ['A3', 'B3', 'C3', 'D3', 'E3', 'F3']:
            ws[t].border = myborder
            ws[t].fill = white
        ws.merge_cells('A3:F3')

        ws['A4'].value = '3.6V'
        ws['A4'].alignment = ali
        ws['A4'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (4, 14):
            ws['A'+str(t)].border = myborder
            ws['A' + str(t)].fill = SpringGreen2
            ws['B' + str(t) ].fill = SpringGreen2
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A4:A13')

        ws['A14'].value = '3.8V'
        ws['A14'].alignment = ali
        ws['A14'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (14, 24):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A14:A23')

        ws['A24'].value = '4.0V'
        ws['A24'].alignment = ali
        ws['A24'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (24, 34):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Turquoise
            ws['B' + str(t)].fill = Turquoise
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A24:A33')

        ws['A34'].value = '4.1V'
        ws['A34'].alignment = ali
        ws['A34'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (34, 44):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Yellow3
            ws['B' + str(t)].fill = Yellow3
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A34:A43')

        ws['A44'].value = '4.2V'
        ws['A44'].alignment = ali
        ws['A44'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (44, 54):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A44:A53')

        # 第二轮老化测试
        ws['A54'].value = '第二轮老化测试'
        ws['A54'].alignment = ali
        ws['A54'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in ['A54', 'B54', 'C54', 'D54', 'E54', 'F54']:
            ws[t].border = myborder
            ws[t].fill = white
        ws.merge_cells('A54:F54')

        ws['A55'].value = '3.6V'
        ws['A55'].alignment = ali
        ws['A55'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (55, 65):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = SpringGreen2
            ws['B'+ str(t)].fill = SpringGreen2
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A55:A64')

        ws['A65'].value = '3.8V'
        ws['A65'].alignment = ali
        ws['A65'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (65, 75):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A65:A74')

        ws['A75'].value = '4.0V'
        ws['A75'].alignment = ali
        ws['A75'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (75, 85):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Turquoise
            ws['B' + str(t)].fill = Turquoise
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A75:A84')

        ws['A85'].value = '4.1V'
        ws['A85'].alignment = ali
        ws['A85'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (85, 95):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Yellow3
            ws['B' + str(t)].fill = Yellow3
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A85:A94')

        ws['A95'].value = '4.2V'
        ws['A95'].alignment = ali
        ws['A95'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (95, 105):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A95:A104')

        # 第三轮老化测试
        ws['A105'].value = '第三轮老化测试'
        ws['A105'].alignment = ali
        ws['A105'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in ['A105', 'B105', 'C105', 'D105', 'E105', 'F105']:
            ws[t].border = myborder
            ws[t].fill = white
        ws.merge_cells('A105:F105')

        ws['A106'].value = '3.6V'
        ws['A106'].alignment = ali
        ws['A106'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (106, 116):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = SpringGreen2
            ws['B' + str(t)].fill = SpringGreen2
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A106:A115')

        ws['A116'].value = '3.8V'
        ws['A116'].alignment = ali
        ws['A116'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (116, 126):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].fill = DeepSkyBlue
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A116:A125')

        ws['A126'].value = '4.0V'
        ws['A126'].alignment = ali
        ws['A126'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (126, 136):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Turquoise
            ws['B' + str(t)].fill = Turquoise
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A126:A135')

        ws['A136'].value = '4.1V'
        ws['A136'].alignment = ali
        ws['A136'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (136, 146):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = Yellow3
            ws['B' + str(t)].fill = Yellow3
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A136:A145')

        ws['A146'].value = '4.2V'
        ws['A146'].alignment = ali
        ws['A146'].font = Font(name='黑体', size=11, bold=False, italic=True, color=colors.BLACK)
        for t in (146, 155):
            ws['A' + str(t)].border = myborder
            ws['A' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].fill = PaleGoldenrod
            ws['B' + str(t)].border = myborder
        ws.merge_cells('A146:A155')

        x = range(1, 51)
        y = range(51, 101)
        z = range(101, 151)
        v = range(151, 201)
        q = [q for q in range(1, 11)]
        color_array = [SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,SpringGreen2,
                       DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,DeepSkyBlue,
                       Turquoise, Turquoise, Turquoise, Turquoise, Turquoise,Turquoise,Turquoise,Turquoise,Turquoise,Turquoise,
                       Yellow3, Yellow3,Yellow3,Yellow3,Yellow3,Yellow3,Yellow3,Yellow3,Yellow3,Yellow3,
                       PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod,PaleGoldenrod]

        for i,j,h in zip(range(4, 54), q*5, color_array):
            ws.cell(i, 2, j)
            ws.cell(i,2).fill = h
            ws.cell(i,2).border = myborder

        for i, j,h in zip(range(55, 105), q * 5, color_array):
            ws.cell(i, 2, j)
            ws.cell(i, 2).fill = h
            ws.cell(i, 2).border = myborder

        for i, j,h in zip(range(106, 156), q * 5, color_array):
            ws.cell(i, 2, j)
            ws.cell(i, 2).fill = h
            ws.cell(i, 2).border = myborder

        # 从这里开始开始数据添加
        board_voltage_one = self.board_voltage_[0:50]
        board_voltage_two = self.board_voltage_[50:100]
        board_voltage_three = self.board_voltage_[100:]

        board_current_one = self.board_current_[0:50]
        board_current_two = self.board_current_[50:100]
        board_current_three = self.board_current_[100:]

        mu_voltage_one = self.mu_voltage_[0:50]
        mu_voltage_two = self.mu_voltage_[50:100]
        mu_voltage_three = self.mu_voltage_[100:]

        mu_current_one = self.mu_current_[0:50]
        mu_current_two = self.mu_current_[50:100]
        mu_current_three = self.mu_current_[100:]

        # 第一组数据添加
        for j,i in zip(range(4, 54), board_voltage_one):
            ws.cell(j, 3, i)

        for j,i in zip(range(4, 54),mu_voltage_one):
            ws.cell(j, 4, i)

        for j, i in zip(range(4, 54), board_current_one):
            ws.cell(j, 5, i)

        for j, i in zip(range(4, 54), mu_current_one):
            ws.cell(j, 6, i)

        # 第二组数据添加
        for j,i in zip(range(55, 105), board_voltage_two):
            ws.cell(j, 3, i)

        for j,i in zip(range(55, 105),mu_voltage_two):
            ws.cell(j, 4, i)

        for j, i in zip(range(55, 105), board_current_two):
            ws.cell(j, 5, i)

        for j, i in zip(range(55, 105), mu_current_two):
            ws.cell(j, 6, i)


        # 第三组数据添加
        for j,i in zip(range(106, 156), board_voltage_three):
            ws.cell(j, 3, i)

        for j,i in zip(range(106, 156), mu_voltage_three):
            ws.cell(j, 4, i)

        for j, i in zip(range(106, 156), board_current_three):
            ws.cell(j, 5, i)

        for j, i in zip(range(106, 156), mu_current_three):
            ws.cell(j, 6, i)

        prev_msg = str(time.localtime().tm_year) + '_' + str(time.localtime().tm_mon) + '_' + str(
            time.localtime().tm_mday) + '_' + str(time.localtime().tm_hour) + '_' + str(
            time.localtime().tm_min) + '_' + str(
            time.localtime().tm_sec)
        if self.board_num_ is None:
            middle_msg = 'sheet1'
        else:
            middle_msg = str(self.board_num_)
        next_msg = '1000r'
        msg = prev_msg+'_'+middle_msg+'_'+next_msg
        wb.save('./'+msg+'.xlsx')


def main():
    e = Excel_Template([1],[1], [1], [1], '1')
    # print((1 and 1 and 1 and 1 and 1) is False)

if __name__ == '__main__':
    main()