# -*- coding:utf-8 -*-
# 导入字体,边框,颜色以及对齐的方式
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
import openpyxl


def main():
    wb = openpyxl.Workbook()
    robot_num = str(input('请输入机器编号'))
    # 创建我们自己的sheet
    ws = wb.create_sheet(robot_num, 0)
    # 给A1添加数据
    ws.cell(1, 1, 'Welcome to openpyxl')

    # 设置字体
    # 设置字体风格为Times New Roman，大小为16，粗体、斜体，颜色蓝色
    ws['A1'].font = Font(name='Times New Roman', size=16, bold=True, italic=True, color=colors.BLUE)

    # 设置对齐方式
    # 通过参数horizontal和vertical来设置文字在单元格里的对齐方式，此外设置值还可为left和right
    ws.cell(1, 2, '大家好')
    # 水平和垂直方向都是居中
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    # 设置边框
    # 首先设置边框四个方向的线条种类 thin是实线 颜色是黑色
    left, right, top, bottom = [Side(style='thin', color='000000')] * 4
    # 再将各方向线条作为参数传入Border方法
    ws.cell(1, 3, '我有边框')
    ws['C1'].border = Border(left=left, right=right, top=top, bottom=bottom)


    # 设置行高和列宽
    # 行和列的长度大小可以通过row_dimensions[序号].height和column_dimensions[标号].width来设置
    # 设置第一行的行高
    # 设置行高
    ws.row_dimensions[1].height = 25.0

    # 设置D列宽
    ws.column_dimensions['D'].width = 20.0


    # 合并单元格
    # 对单元格的合并与拆分，主要是通过sheet的merge_cells(args1:args2)和unmerge_cells(args1:args2)两个方法来实现的
    # ws.merge_cells() 有两种使用方式 一种是使用范围字符串
    ws.merge_cells('E1:F1')
    # ws.merge_cells() 有两种使用方式 二种是使用范围数字
    ws.merge_cells(None, 3, 1, 6, 1)
    # 同样的方式拆分合并后的单元格也是两种方式

    # 保存文件
    wb.save('show.xlsx')


if __name__ == '__main__':
    main()